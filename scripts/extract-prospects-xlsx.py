#!/usr/bin/env python3
"""Extrait les prospects d'un Excel de prospection vers JSON (stdout).

Détecte les colonnes par leur entête (1re ligne) : Email, lien d'aperçu Akyra
(« Site Akyra (aperçu) » / contient 'akyra' ou 'aperçu'), « Message à copier-coller »,
Prénom/Nom, Doublon.

    python3 scripts/extract-prospects-xlsx.py "/chemin/fichier.xlsx"
"""
import sys, json, re
import openpyxl

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
TOKEN_RE = re.compile(r"/r/([a-f0-9]+)")


def norm(s):
    return str(s).strip().lower() if s is not None else ""


def find_col(headers, *needles):
    for i, h in enumerate(headers):
        hn = norm(h)
        if any(n in hn for n in needles):
            return i
    return None


def main():
    if len(sys.argv) < 2:
        print("usage: extract-prospects-xlsx.py <path.xlsx>", file=sys.stderr)
        sys.exit(2)
    wb = openpyxl.load_workbook(sys.argv[1], read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        json.dump([], sys.stdout)
        return

    headers = rows[0]
    c_email = find_col(headers, "email")
    c_link = find_col(headers, "aperçu", "apercu", "site akyra")
    c_msg = find_col(headers, "message")
    c_first = find_col(headers, "prénom", "prenom")
    c_nom = find_col(headers, "nom")
    c_dup = find_col(headers, "doublon")

    def cell(row, i):
        return str(row[i]).strip() if i is not None and i < len(row) and row[i] is not None else ""

    out = []
    for row in rows[1:]:
        if not any(c for c in row):
            continue
        email = cell(row, c_email)
        link = cell(row, c_link)
        m = TOKEN_RE.search(link)
        token = m.group(1) if m else None
        first = cell(row, c_first) or cell(row, c_nom)
        dup = cell(row, c_dup)
        out.append({
            "token": token,
            "email": email,
            "first_name": first,
            "message": cell(row, c_msg),
            "doublon": bool(dup),
            "valid": bool(EMAIL_RE.match(email)) and token is not None and not dup,
        })
    json.dump(out, sys.stdout, ensure_ascii=False)


if __name__ == "__main__":
    main()
